"""Export a compact Excel audit of raw and filtered temperature coverage."""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from psmi.data import load_and_prepare_excel  # noqa: E402


DATASETS = [
    (
        "Curated IL-LLE",
        PROJECT_ROOT / "datasets" / "processed" / "update-LLE-all-with-smiles.xlsx",
        "LLE system NO.",
    ),
    (
        "Expanded literature LLE",
        PROJECT_ROOT / "datasets" / "processed" / "LLE-literature-data-boosted.xlsx",
        "System NO.",
    ),
]
OUTPUT = (
    PROJECT_ROOT
    / "experiments"
    / "supporting_information"
    / "s5_dataset_construction_and_distribution"
    / "results"
    / "temperature_range_audit.xlsx"
)


def collect_rows() -> tuple[list[list[object]], list[list[object]]]:
    summary: list[list[object]] = []
    extremes: list[list[object]] = []
    for dataset, path, raw_system_column in DATASETS:
        raw = pd.read_excel(path)
        raw_temperature = pd.to_numeric(raw["T/K"], errors="coerce")
        filtered, _ = load_and_prepare_excel(
            str(path), min_points_per_group=6, permute_23_aug=False
        )
        for stage, frame, temperature, system_column in [
            ("Raw workbook", raw, raw_temperature, raw_system_column),
            ("Filtered: min 6 tie-lines/group", filtered, filtered["T"], "system_id"),
        ]:
            minimum = float(temperature.min())
            maximum = float(temperature.max())
            summary.append(
                [
                    dataset,
                    stage,
                    len(frame),
                    int(pd.Series(frame[system_column]).nunique()),
                    int(temperature.nunique()),
                    minimum,
                    maximum,
                    str(path.relative_to(PROJECT_ROOT)),
                ]
            )
            for label, value in [("Minimum", minimum), ("Maximum", maximum)]:
                mask = temperature.eq(value)
                extremes.append(
                    [
                        dataset,
                        stage,
                        label,
                        value,
                        int(mask.sum()),
                        int(pd.Series(frame.loc[mask, system_column]).nunique()),
                    ]
                )
    return summary, extremes


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def main() -> None:
    summary, extremes = collect_rows()
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Temperature range"
    overview.append(
        [
            "Dataset",
            "Stage",
            "Records",
            "Unique systems",
            "Unique temperatures",
            "Minimum T (K)",
            "Maximum T (K)",
            "Source workbook",
        ]
    )
    for row in summary:
        overview.append(row)
    support = workbook.create_sheet("Extreme-value support")
    support.append(
        ["Dataset", "Stage", "Extreme", "Temperature (K)", "Records", "Systems"]
    )
    for row in extremes:
        support.append(row)
    for sheet in [overview, support]:
        style_sheet(sheet)
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 32
    overview.column_dimensions["C"].width = 14
    overview.column_dimensions["D"].width = 16
    overview.column_dimensions["E"].width = 20
    overview.column_dimensions["F"].width = 17
    overview.column_dimensions["G"].width = 17
    overview.column_dimensions["H"].width = 62
    for column in "ABCDEF":
        support.column_dimensions[column].width = 25
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    verified = load_workbook(OUTPUT, data_only=False, read_only=True)
    assert verified.sheetnames == ["Temperature range", "Extreme-value support"]
    print(OUTPUT)


if __name__ == "__main__":
    main()
