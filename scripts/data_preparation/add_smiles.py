"""Add component SMILES columns to curated LLE workbooks."""

from pathlib import Path
from typing import Dict, Optional
from openpyxl import load_workbook, Workbook

# Configurable file names
PROJECT_ROOT = Path(__file__).resolve().parents[2]
INPUT_LLE_XLSX = PROJECT_ROOT / "datasets" / "raw" / "AIChEj-LLE-all.xlsx"
MAPPING_XLSX = PROJECT_ROOT / "datasets" / "raw" / "更新-LLE_components-手动查找smiles.xlsx"
OUTPUT_XLSX = PROJECT_ROOT / "datasets" / "processed" / "update-LLE-all-with-smiles.xlsx"

# Expected component column headers in the LLE sheet
COMPONENT_COLS = [
    "IL (Component 1) full name",
    "Component 2",
    "Component 3",
]


def load_name_to_smiles_from_excel(path: Path) -> Dict[str, str]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    # Identify header row
    header = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Find name and smiles columns
    name_idx: Optional[int] = None
    smiles_idx: Optional[int] = None
    for i, h in enumerate(header):
        hh = str(h).strip().lower()
        if hh in ("name", "component", "compound", "组分"):
            if name_idx is None:
                name_idx = i
        if "smiles" in hh:
            smiles_idx = i
    if name_idx is None:
        # fallback: first column
        name_idx = 0
    if smiles_idx is None:
        # try second column fallback
        smiles_idx = 1 if len(header) > 1 else 0
    mapping: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        name = row[name_idx].value
        smiles = row[smiles_idx].value
        if name is None:
            continue
        key = str(name).strip()
        val = "" if smiles is None else str(smiles).strip()
        if key:
            mapping[key] = val
    wb.close()
    return mapping


def add_smiles_columns():
    if not INPUT_LLE_XLSX.exists():
        print(f"Input LLE Excel not found: {INPUT_LLE_XLSX}")
        return
    if not MAPPING_XLSX.exists():
        print(f"Mapping Excel not found: {MAPPING_XLSX}")
        return
    name_to_smiles = load_name_to_smiles_from_excel(MAPPING_XLSX)
    wb = load_workbook(filename=INPUT_LLE_XLSX)
    ws = wb.active

    # Read header to find component columns and their indices
    header_cells = list(ws.iter_rows(min_row=1, max_row=1))[0]
    headers = [str(c.value) if c.value is not None else "" for c in header_cells]

    comp_indices = []
    for comp in COMPONENT_COLS:
        try:
            idx = headers.index(comp)
            comp_indices.append(idx)
        except ValueError:
            comp_indices.append(None)

    # Insert new SMILES columns immediately to the right of each component column
    # Need to handle shifting indices when inserting; process from rightmost to leftmost
    insertion_plan = []
    for label, idx in zip(COMPONENT_COLS, comp_indices):
        if idx is not None:
            insertion_plan.append((label, idx))
    insertion_plan.sort(key=lambda x: x[1], reverse=True)

    for label, idx in insertion_plan:
        ws.insert_cols(idx + 2)  # openpyxl is 1-based, and we want right of idx -> +2
        ws.cell(row=1, column=idx + 2, value=f"{label} SMILES")

    # Build updated header list after inserts
    header_cells = list(ws.iter_rows(min_row=1, max_row=1))[0]
    headers = [str(c.value) if c.value is not None else "" for c in header_cells]

    # Map each component column to its SMILES column index
    col_map = {}
    for label in COMPONENT_COLS:
        try:
            ci = headers.index(label)
            si = headers.index(f"{label} SMILES")
            col_map[label] = (ci, si)
        except ValueError:
            pass

    # Fill SMILES values for each row
    for row in ws.iter_rows(min_row=2):
        for label, (ci, si) in col_map.items():
            name_val = row[ci].value
            name_str = "" if name_val is None else str(name_val).strip()
            smiles = name_to_smiles.get(name_str, "")
            row[si].value = smiles

    wb.save(OUTPUT_XLSX)
    wb.close()
    print(f"Wrote updated Excel with SMILES: {OUTPUT_XLSX}")


if __name__ == "__main__":
    add_smiles_columns()
